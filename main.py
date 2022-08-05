
##    AUTOMATION WITH PYTHON
## Learn how to work with spreadsheets
## Read spreadsheet and automate stuff

##      EXERCISES
## EX1: List each company with respective product count
## EX3: List products with inventory less than 10
## EX2: List each company with respective total inventory value
## Ex4 : Write to spreadsheet : Calculate and write inventory value for each product into spreadsheet

## Python has several built in functions in general e.g IO module : create, read, write
## Python has a package to work with spreadsheet  specifically called OPENPYXL

import openpyxl  ## importing the package into this file after installing

inv_file = openpyxl.load_workbook("inventory.xlsx") ## get the content of the inventory file
product_list = inv_file["Sheet1"] ## get all information from the sheet1

##  Exercile 1 : Calculate how many productss we have per supplier and list the supplier with  that respective number of products
## we are going to loop, so we go through each and every row in the sheet and also execute same logic on each item
## we are going through as many times of the product i.e 74

product_per_supplier = {}  ## creating an empty dictionary  get the number of product per supplier
total_value_per_supplier = {} ## creating an empty dictonary that get the total value of suppliers
product_sunder_10_inv = {}  ## creating an empty dictionary for product that has inventory less than 10
# print(product_list.max_row)  ## gets the number of rows in the worksheet

# for product_row in range(product_list.max_row): ## we using a range because it will create a sequence of numbers starting from by default. we can then now iterate over and do something for each item in the list [0,1,2,3 ....]
for product_row in range(2, product_list.max_row + 1): ## The 2 gets the scond row and +1 gets the last row
    supplier_name = product_list.cell(product_row, 4).value ## this will give us all suppliers values for each row, because the supplier values is on line 4
    inventory = product_list.cell(product_row, 2).value   ## this will give us all inventory values for each row, because the inventory values is on line 2
    price = product_list.cell(product_row, 3).value   ## this will give us all price values for each row, because the price values is on line 3
    product_num = product_list.cell(product_row, 1).value   ## this will give us all product values for each row, because the product values is on line 1
    inventory_price = product_list.cell(product_row, 5) ## this will give us all empty values for each row, because the empty values is on line 5
   
    ## calculation of total number of product per supplier
    if supplier_name in product_per_supplier: ## this means if the supplier_name is in the dictionary lsit
        current_num_products = product_per_supplier.get(supplier_name) ## getting the value from the dictionary using the key name
        product_per_supplier[supplier_name] = current_num_products + 1 ## incrementing the value gotten
    else:
        product_per_supplier[supplier_name] = 1


## Calculation of total value of inventory per supplier

    if supplier_name in total_value_per_supplier: ## if the supplier_name is present in the total_value_per_supplier then it should just add the values to the existing value
        current_total_value = total_value_per_supplier.get(supplier_name) ## getting the current tota; value
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price ## multiplying the values of both inventory and price and then getting the total value

## EX3 printing out product that has inventory less than 10   
#   Logic:    product with inventory less than 10     
    if inventory < 10:
        product_sunder_10_inv[int(product_num)] = int(inventory)
    

# Ex4 : Write to spreadsheet : Calculate and write inventory value for each product into spreadsheet
#    add value for total inventory price
    inventory_price.value = inventory * price ## creating a new value for the 5th column and getting the value
print(product_per_supplier)   ##  printing out the companies name(key) and value
print(total_value_per_supplier) ## getting the total values of supplier for each company by multiplying the price and inventory of each company and adding the total values together
print(product_sunder_10_inv) ## getting the products with less than 10 inventory

inv_file.save("inventory_with_total_value.xlsx")  ##  create a new file and save with the name in the string
